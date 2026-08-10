from __future__ import annotations

import ast
import json
import re
import sys
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
APP = ROOT / "app.py"
CLEAN = ROOT / "field_trial_data_clean_seed.xlsx"

results: list[dict] = []

def check(name: str, passed: bool, detail: str = "") -> None:
    results.append({"check": name, "passed": bool(passed), "detail": detail})

source = APP.read_text(encoding="utf-8")
component_source = (ROOT / "photo_component" / "index.html").read_text(encoding="utf-8")
photo_integrity_source = (ROOT / "photo_integrity.py").read_text(encoding="utf-8")
combined_source = source + "\n" + component_source + "\n" + photo_integrity_source

try:
    ast.parse(source)
    check("Python syntax", True)
except SyntaxError as exc:
    check("Python syntax", False, str(exc))

banned = {
    "Embedded camera removed": "st.camera_input(",
    "Old route wording removed": "Route point ",
    "Zoom remains enabled": "user-scalable=no",
    "No maximum scale lock": "maximum-scale",
    "No sticky CTA selector": ".element-container:has(.mobile-save-anchor)",
    "No brittle checked-card selector": ":has(.checked-trap-marker)",
    "No floating saved banner": "saved-check-banner",
}
for name, token in banned.items():
    check(name, token not in source, f"Forbidden token: {token}")

required = {
    "Browser photo component": "PHOTO_COMPONENT = components.declare_component",
    "Multiple image upload": "multiple />",
    "Browser-side resize": "Math.min(1800, sourceEdge)",
    "Browser-side JPEG quality": "qualities=[0.82",
    "Automatic retry backoff": "retry_delays_ms=[1000, 2000, 4000]",
    "Stable pending Check ID": "def ensure_pending_check_id(",
    "Idempotent photo path": "def _pending_image_path(",
    "Final photo verification": "def verify_pending(",
    "Central navigation controller": "def navigate(page: str, rerun: bool = True, **kwargs):",
    "Top-reset navigation": "st.session_state.scroll_to_top_once = True",
    "App-owned top anchor": "r1m1-page-top",
    "Delayed scroll retries": "[80, 200, 450, 900]",
    "Card wrapper": "def app_card():",
    "Stable visit card component": "def render_visit_trap_card(",
    "Checked card app class": "visit-trap-card is-checked",
    "Stateful data navigation": 'key="data_management_section"',
    "Safe-area handling": "env(safe-area-inset-bottom)",
    "Mobile 16px inputs": "font-size: 16px !important",
    "Radio styling selector": 'label[data-baseweb="radio"]',
    "Radio unselected white surface": 'background: #ffffff !important;',
    "Radio unselected dark outline": 'border: 2px solid #444a53 !important;',
    "Radio selected orange outline": 'border-color: var(--brand-orange) !important;',
    "Radio selected orange centre": 'background: var(--brand-orange) !important;',
    "Checkbox styling": 'label[data-baseweb="checkbox"]',
    "Hidden framework navigation router": 'selected_navigation_page = st.navigation(NAVIGATION_PAGES, position="hidden")',
    "Wrapping primary navigation": 'with st.container(',
    "Native page links": "st.page_link(",
    "Administration popover": 'with st.popover("Administration", key=f"app_top_navigation_admin_popover_',
    "Top navigation administration group": '"Administration": [',
    "Top navigation sign out": 'st.Page(top_nav_sign_out, title="Sign out"',
    "Traps page title": 'header("Traps",',
    "Follow-ups page title": 'header("Follow-ups",',
    "Trial performance page title": 'header("Trial performance",',
    "Trial setup page title": 'header("Trial setup",',
    "Data and records page title": 'header("Data & records",',
    "Trial periods page title": 'header("Trial periods",',
    "Grouped trap-history day heading": 'class="trap-history-day"',
    "Fixed history time column": "grid-template-columns: 5.25rem minmax(0, 1fr);",
    "Trap editor close control": 'key="close_trap_setup_panel_top"',
}
for name, token in required.items():
    check(name, token in combined_source, f"Required token: {token}")

wb = load_workbook(CLEAN, data_only=False, read_only=True)
sheets = ["Sites","Builds","Traps","Visits","Checks","Windows","Followups","Audit Log","Photos"]
data = {}

missing_sheets = [name for name in sheets if name not in wb.sheetnames]
check("Required workbook sheets", not missing_sheets, json.dumps(missing_sheets))

for sheet_name in sheets:
    if sheet_name not in wb.sheetnames:
        data[sheet_name] = []
        continue
    ws = wb[sheet_name]
    values = list(ws.iter_rows(min_row=1, max_row=1000, max_col=52, values_only=True))
    if not values:
        data[sheet_name] = []
        continue
    headers = [str(v) if v is not None else "" for v in values[0]]
    rows = []
    for row in values[1:]:
        if any(v not in (None, "") for v in row):
            padded = list(row) + [None] * max(0, len(headers) - len(row))
            rows.append(dict(zip(headers, padded[:len(headers)])))
    data[sheet_name] = rows

expected_counts = {
    "Sites": 3, "Traps": 15, "Visits": 0, "Checks": 0,
    "Windows": 15, "Followups": 0, "Audit Log": 0, "Photos": 0
}
actual_counts = {name: len(data[name]) for name in expected_counts}
check("Clean seed counts", actual_counts == expected_counts, json.dumps(actual_counts))

def values(rows, key):
    return [str(r.get(key)) for r in rows if r.get(key) not in (None, "")]

def duplicates(items):
    return sorted([k for k, v in Counter(items).items() if v > 1])

for sheet_name, key in [("Sites","Site ID"),("Traps","Trap ID"),("Windows","Window ID")]:
    dups = duplicates(values(data[sheet_name], key))
    check(f"Unique {sheet_name} IDs", not dups, json.dumps(dups))

site_ids = set(values(data["Sites"], "Site ID"))
trap_ids = set(values(data["Traps"], "Trap ID"))
broken = []
for row in data["Traps"]:
    if str(row.get("Site ID")) not in site_ids:
        broken.append(f"Trap {row.get('Trap ID')} → {row.get('Site ID')}")
for row in data["Windows"]:
    if str(row.get("Site ID")) not in site_ids:
        broken.append(f"Window {row.get('Window ID')} site")
    if str(row.get("Trap ID")) not in trap_ids:
        broken.append(f"Window {row.get('Window ID')} trap")
check("Workbook references", not broken, json.dumps(broken))

open_counts = Counter(
    str(row.get("Trap ID"))
    for row in data["Windows"]
    if str(row.get("Status")) == "Open"
)
bad_open = {tid: open_counts.get(tid, 0) for tid in trap_ids if open_counts.get(tid, 0) != 1}
check("One open window per trap", not bad_open, json.dumps(bad_open))

formula_errors = []
error_pattern = re.compile(r"#REF!|#DIV/0!|#VALUE!|#NAME\\?|#N/A")
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for cell in row:
            value = cell.value
            if isinstance(value, str) and error_pattern.search(value):
                formula_errors.append(f"{ws.title}!{cell.coordinate}: {value}")
                if len(formula_errors) >= 300:
                    break
        if len(formula_errors) >= 300:
            break
    if len(formula_errors) >= 300:
        break
check("Workbook formula errors", not formula_errors, json.dumps(formula_errors[:20]))

failed = [item for item in results if not item["passed"]]
print(json.dumps({
    "passed": not failed,
    "checks": results,
    "failed_count": len(failed),
}, indent=2))
sys.exit(1 if failed else 0)


# v8.6.53 architecture markers
